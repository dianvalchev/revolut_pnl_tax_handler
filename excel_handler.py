import datetime
import openpyxl as xl

class ExcelHandler:

    def __init__(self, xl_file):
        self.result_file = f"ResultFile {datetime.datetime.strftime(datetime.datetime.now(),'%Y-%m-%d %Hh%Mm%Ss')}.xlsx"
        self.header_row = None
        self.xl_file = xl_file
        self.data = self.open_pnl_file()

    def open_pnl_file(self):
        # load the pnl file with openpyxl
        wb = xl.load_workbook(self.xl_file)
        ws = wb.active

        # store the header row in a list
        self.header_row = [x.value for x in ws[2]]

        # store the remaining rows in a list
        data = list(ws.iter_rows(min_row=3, values_only=True))

        return data

    def create_result_file(self):
        # create the file and delete the default sheet
        wb = xl.Workbook()
        ws = wb.active
        wb.remove(ws)

        # parse the data list and create a new sheet for every year in the list
        for item in self.data:
            if item[1] is None:
                break

            sheet_name = str(item[1].year)
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
                wb[sheet_name].append(self.header_row)
            # add the row to the correct sheet
            wb[sheet_name].append(item)

        # sort the sheets by year and save the file
        wb._sheets.sort(key=lambda sheet: sheet.title)
        wb.save(self.result_file)

    def add_bgn_pnl_to_result_file(self, data):
        # load the created results file
        wb = xl.load_workbook(self.result_file)

        # parse all the sheets and fill BGN data
        for sheet in wb.sheetnames:
            ws = wb[sheet]

            # add the new header row items (BGN)
            for item in ['Date Acquired (BGN/USD)', 'Date Sold (BGN/USD)', 'Cost basis (BGN)', 'Amount (BGN)',
                         'Realised PnL (BGN)']:
                ws.cell(row=1, column=ws.max_column + 1).value = item

            # calculate and fill the PnL data in BGN
            for row in range(2, ws.max_row + 1):
                date_acquired = ws.cell(row=row, column=1).value.strftime('%Y-%m-%d')
                date_sold = ws.cell(row=row, column=2).value.strftime('%Y-%m-%d')

                # get the fx rates for the selected dates
                ws.cell(row=row, column=12).value = data[date_acquired]
                ws.cell(row=row, column=13).value = data[date_sold]

                # fill the BGN amounts in corresponding cells
                ws.cell(row=row, column=14).value = f'=H{row}*L{row}' # Cost basis (BGN)
                ws.cell(row=row, column=15).value = f'=I{row}*M{row}' # Amount (BGN)
                ws.cell(row=row, column=16).value = f'=O{row}-N{row}' # Realised PnL (BGN)

                # format the cells in the corresponding style
                ws.cell(row=row, column=1).number_format = "YYYY.MM.DD"
                ws.cell(row=row, column=2).number_format = "YYYY.MM.DD"
                ws.cell(row=row, column=8).number_format = "$ #,##0.00"
                ws.cell(row=row, column=9).number_format = "$ #,##0.00"
                ws.cell(row=row, column=10).number_format = "$ #,##0.00"
                ws.cell(row=row, column=12).number_format = "#,##0.00 лв."
                ws.cell(row=row, column=13).number_format = "#,##0.00 лв."
                ws.cell(row=row, column=14).number_format = "#,##0.00 лв."
                ws.cell(row=row, column=15).number_format = "#,##0.00 лв."
                ws.cell(row=row, column=16).number_format = "#,##0.00 лв."

            # fill the total due amount at the end of the sheet
            ws.cell(row=ws.max_row + 1, column=16).value = f'=SUM(P2:P{ws.max_row})'
            ws.cell(row=ws.max_row, column=16).number_format = "#,##0.00 лв."
        wb.save(self.result_file)


# Test the Excel handler
if __name__ == "__main__":
    example_data = {'2019-11-27': 1.77657, '2020-01-28': 1.77722, '2019-11-25': 1.77674, '2021-09-14': 1.65552,
                    '2022-06-10': 1.84896, '2021-11-15': 1.70904, '2020-02-28': 1.78175, '2020-06-05': 1.72624,
                    '2021-08-16': 1.66143, '2022-02-07': 1.7086, '2020-01-29': 1.77787, '2019-12-06': 1.76296,
                    '2021-01-22': 1.60868, '2020-07-02': 1.73297, '2020-07-06': 1.727, '2021-07-09': 1.64938,
                    '2020-08-04': 1.66241, '2020-04-14': 1.78403, '2019-12-12': 1.75616, '2020-03-12': 1.74006,
                    '2020-01-17': 1.76074, '2023-10-16': 1.85598, '2019-11-26': 1.7748, '2023-03-02': 1.84425,
                    '2020-01-03': 1.75458, '2020-04-06': 1.81246, '2020-10-08': 1.66241, '2021-10-20': 1.68272,
                    '2019-12-02': 1.77432, '2022-09-07': 1.97858, '2023-09-05': 1.8226, '2021-05-25': 1.59477,
                    '2020-11-02': 1.67854, '2021-02-16': 1.61066, '2021-03-29': 1.65973, '2022-03-03': 1.76106,
                    '2020-05-01': 1.7983,'2020-07-15': 1.7983,'2024-12-18': 1.7983, '2020-08-03': 1.7983,
                    '2020-12-01': 1.7983,'2021-01-08': 1.7983,'2022-01-27': 1.7983,'2022-01-24': 1.7983,
                    '2020-11-26': 1.7983,'2021-11-27': 1.7983,'2022-11-27': 1.7983,'2024-10-14': 1.7983,
                    '2021-02-03': 1.7983,'2022-03-04': 1.7983}

    statement_file_name = "trading-pnl-statement_2019-11-24_2025-01-22_en-us_bf01c9.xlsx"  # 2019-2023 PnL
    my_handler = ExcelHandler(statement_file_name)
    my_handler.create_result_file()
    my_handler.add_bgn_pnl_to_result_file(example_data)
